# -*- coding: utf-8 -*-
"""SM 촬영수(sm_shoot_daily.parquet) → 부서 공유용 엑셀(아티스트별 탭).

현재 오픈 중인 IP만, 한 엑셀에 아티스트별 시트로 나눠 보기 좋게 만든다.
멤버·테마의 한/영 표기를 한국어로 통합. 값 = 촬영수(Artist별 촬영수, CMS와 일치).

시트: 요약 → (아티스트별) → 국가별
실행:
  python sm_report.py                      # 전체 기간 → reports/
  python sm_report.py 2026-06-16 2026-06-29
"""
import datetime as dt
import io
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

BASE_DIR = Path(__file__).parent
DAILY_PARQUET = BASE_DIR / "data" / "sm_shoot_daily.parquet"
CHANGES_PARQUET = BASE_DIR / "data" / "sm_shoot_changes.parquet"  # 주차별 변동 이력
CONFIG_FILE = BASE_DIR / "config.json"
REPORT_DIR = BASE_DIR / "reports"

# ── 현재 오픈 IP (순서 = 시트 순서). kws=테마 부분일치(소문자), members=한국어 통합명→별칭 ──
ARTISTS = [
    {"name": "NCT WISH", "kws": ["nct wish"], "countries": None, "members": {
        "시온": ["시온", "sion"], "리쿠": ["리쿠", "riku"], "유우시": ["유우시", "유시", "yushi"],
        "사쿠야": ["사쿠야", "sakuya"], "재희": ["재희", "jaehee"], "료": ["료", "ryo"]}},
    {"name": "라이즈", "kws": ["riize", "라이즈"], "countries": None, "members": {
        "쇼타로": ["쇼타로", "shotaro"], "은석": ["은석", "eunseok"], "성찬": ["성찬", "sungchan"],
        "원빈": ["원빈", "wonbin"], "앤톤": ["앤톤", "안톤", "anton"], "소희": ["소희", "sohee"]}},
    {"name": "에스파", "kws": ["aespa", "에스파"], "countries": None, "members": {
        "카리나": ["카리나", "karina"], "지젤": ["지젤", "giselle"],
        "윈터": ["윈터", "winter"], "닝닝": ["닝닝", "ningning"]}},
    {"name": "아이린", "kws": ["irene", "아이린"], "countries": None, "members": {
        "아이린": ["아이린", "irene"]}},
    {"name": "승한", "kws": ["xnghan", "승한", "seunghan"], "countries": None, "members": {
        "승한": ["승한", "xnghan", "seunghan"]}},
    {"name": "태용", "kws": ["taeyong", "태용"], "countries": None, "members": {
        "태용": ["태용", "taeyong"]}},
    {"name": "샤이니", "kws": ["shinee", "샤이니"], "countries": None, "members": {
        "온유": ["온유", "onew"], "키": ["키", "key"], "민호": ["민호", "minho"], "태민": ["태민", "taemin"]}},
    {"name": "NCT 재민제노", "kws": ["jnj", "재민제노"], "countries": ["jp"], "members": {
        "재민": ["재민", "jaemin"], "제노": ["제노", "jeno"], "재민제노(듀오)": ["jnjm", "재민제노"]}},
    {"name": "려욱", "kws": ["ryeowook", "려욱"], "countries": None, "members": {
        "려욱": ["려욱", "ryeowook"]}},
]

# ── 서식 ──
# 색은 단순하게 — 회색 2톤만(헤더/합계). 변동(빨강)·증감 색만 기능상 유지.
_HEAD_FILL = PatternFill("solid", fgColor="E5E8EE")   # 헤더: 연회색
_GRAND = PatternFill("solid", fgColor="D5DAE3")       # 전체 합계: 살짝 진한 회색
_TOTAL = PatternFill("solid", fgColor="D5DAE3")       # 국가 소계: 동일 회색
_LIGHT = None                                          # (미사용)
_WHITE_BOLD = Font(bold=True, color="FFFFFF", name="맑은 고딕")
_HEAD_FONT = Font(bold=True, color="33373F", name="맑은 고딕")  # 진한 회색
_CHG_FONT = Font(color="C0392B", name="맑은 고딕")  # 변동 셀(빨강) — 메모 동반
_BOLD = Font(bold=True, name="맑은 고딕")
_NORM = Font(name="맑은 고딕")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_NUMFMT = '#,##0'   # 0은 그대로 '0'으로 표기(담당자 요청)
_thin = Side(style="thin", color="D7DEEE")
_MED = Side(style="medium", color="8AA4E0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEADBORD = Border(left=_thin, right=_thin, top=_MED, bottom=_MED)


def _fit_page(ws):
    """인쇄 시 가로로 안 잘리게 가로방향 + 너비 1페이지 맞춤."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def _mmdd(d) -> str:
    """'2026-06-16' → '06/16' (Excel 날짜서식 미적용 문제 회피, 텍스트로 표기)."""
    s = str(d)
    return f"{s[5:7]}/{s[8:10]}" if len(s) >= 10 and s[4] == "-" else s


def _ctry_label(code, name) -> str:
    """'한국 (KR)' — 한글명 + ISO 국가코드 병기(표기 표준화)."""
    return f"{name} ({str(code).upper()})"


def country_name_map() -> dict:
    try:
        cfg = json.load(open(CONFIG_FILE, encoding="utf-8"))["photoism"]["countries"]
        return {cc: info.get("name", cc.upper()) for cc, info in cfg.items()}
    except Exception:
        return {}


def load_daily(start=None, end=None) -> pd.DataFrame:
    df = pd.read_parquet(DAILY_PARQUET)
    if start:
        df = df[df["날짜"].astype(str) >= str(start)]
    if end:
        df = df[df["날짜"].astype(str) <= str(end)]
    nm = country_name_map()
    df = df.copy()
    df["국가"] = df["국가코드"].map(lambda c: nm.get(c, c.upper()))
    return df


def _match_artist(theme: str, cc: str):
    tl = str(theme).lower()
    for a in ARTISTS:
        if any(k in tl for k in a["kws"]):
            if a["countries"] and cc not in a["countries"]:
                return None
            return a
    return None


def _canon_member(artist, frame: str) -> str:
    fl = str(frame).strip().lower()
    for canon, aliases in artist["members"].items():
        if fl in [x.lower() for x in aliases]:
            return canon
    return str(frame).strip()  # 미정의 멤버는 원문 유지


def annotate(df: pd.DataFrame) -> pd.DataFrame:
    """오픈 아티스트만 남기고 아티스트·멤버(한국어 통합) 컬럼 부여."""
    arts, mems = [], []
    for theme, cc, frame in zip(df["테마"], df["국가코드"], df["프레임"]):
        a = _match_artist(theme, cc)
        if a:
            arts.append(a["name"])
            mems.append(_canon_member(a, frame))
        else:
            arts.append(None)
            mems.append(None)
    out = df.copy()
    out["아티스트"] = arts
    out["멤버"] = mems
    return out[out["아티스트"].notna()]


def aggregate_members(df: pd.DataFrame) -> pd.DataFrame:
    """리포트 셀 단위(날짜·국가·아티스트·멤버) 촬영수 — 주차별 변동 비교용."""
    a = annotate(df)
    if a.empty:
        return pd.DataFrame(columns=["날짜", "국가코드", "아티스트", "멤버", "촬영수"])
    return a.groupby(["날짜", "국가코드", "아티스트", "멤버"], as_index=False)["촬영수"].sum()


def detect_unmatched(df: pd.DataFrame, recent_days: int = 7) -> pd.DataFrame:
    """ARTISTS에 없는 SM 테마 중 '최근 판매중'인 것 = 신규/누락 IP 후보.
    반환: 테마·프레임수·최근판매·마지막판매일·총촬영수 (없으면 빈 DF)."""
    if df.empty:
        return pd.DataFrame()
    matched = annotate(df)
    mk = set(zip(matched["날짜"], matched["국가코드"], matched["테마"], matched["프레임"]))
    idx = df.set_index(["날짜", "국가코드", "테마", "프레임"]).index
    unm = df[~idx.isin(mk)].copy()
    if unm.empty:
        return pd.DataFrame()
    days = sorted(df["날짜"].astype(str).unique())
    recent_set = set(days[-recent_days:])
    g = unm.groupby("테마").agg(프레임수=("프레임", "nunique"),
                                 총촬영수=("촬영수", "sum")).reset_index()
    rec = unm[unm["날짜"].astype(str).isin(recent_set)].groupby("테마")["촬영수"].sum()
    last = unm[unm["촬영수"] > 0].groupby("테마")["날짜"].max()
    g["최근판매"] = g["테마"].map(rec).fillna(0).astype(int)
    g["마지막판매"] = g["테마"].map(last)
    g["총촬영수"] = g["총촬영수"].astype(int)
    return g[g["최근판매"] > 0].sort_values("최근판매", ascending=False).reset_index(drop=True)


def load_changes() -> pd.DataFrame:
    try:
        return pd.read_parquet(CHANGES_PARQUET)
    except Exception:
        return pd.DataFrame(columns=["갱신일", "날짜", "국가코드", "아티스트", "멤버", "이전", "신규"])


def _changes_index(ch: pd.DataFrame) -> dict:
    """(날짜, 국가코드, 아티스트, 멤버) → (이전, 신규, 갱신일) — 셀별 최신 변동."""
    idx = {}
    if ch.empty:
        return idx
    ch = ch.sort_values("갱신일")  # 오름차순 → 마지막(최신)이 최종 반영
    for upd, day, cc, art, mem, prev, new in zip(
            ch["갱신일"], ch["날짜"], ch["국가코드"], ch["아티스트"],
            ch["멤버"], ch["이전"], ch["신규"]):
        idx[(str(day), str(cc), str(art), str(mem))] = (int(prev), int(new), str(upd))
    return idx


def _member_order(name: str, present: set):
    """정의된 전체 멤버(출력 0이어도 표기) + 데이터에만 있는 멤버 뒤에 추가."""
    for a in ARTISTS:
        if a["name"] == name:
            defined = list(a["members"].keys())          # 0건이어도 모두 표기
            extra = sorted(present - set(defined))
            return defined + extra
    return sorted(present)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(2, c)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = _CENTER
        cell.border = _HEADBORD


def _write_pivot(ws, pivot: pd.DataFrame, title: str, row_label: str, date_cols: bool = True):
    """pivot: index=행라벨, columns=날짜(or 아티스트/국가), 값=정수. 렌더+합계.
    date_cols=False면 컬럼 헤더를 날짜서식 없이 원문으로."""
    dates = list(pivot.columns)
    ncols = 1 + len(dates) + 1  # 라벨 + 컬럼들 + 합계
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=13, name="맑은 고딕", color="1A1A2E")
    t.alignment = _LEFT
    # 헤더
    ws.cell(2, 1, row_label)
    for j, d in enumerate(dates):
        ws.cell(2, 2 + j, _mmdd(d) if date_cols else str(d))
    ws.cell(2, ncols, "합계")
    _style_header(ws, ncols)

    col_tot = [int(pivot[d].sum()) for d in dates]
    # 합계 행을 상단(3행)에 — 전체부터 보이게
    hc = ws.cell(3, 1, "합계")
    hc.font, hc.fill, hc.alignment, hc.border = _HEAD_FONT, _GRAND, _LEFT, _BORDER
    for j in range(len(dates)):
        cell = ws.cell(3, 2 + j, col_tot[j])
        cell.number_format = _NUMFMT
        cell.font, cell.fill, cell.alignment, cell.border = _HEAD_FONT, _GRAND, _CENTER, _BORDER
    g = ws.cell(3, ncols, sum(col_tot))
    g.number_format = _NUMFMT
    g.font, g.fill, g.alignment, g.border = _HEAD_FONT, _GRAND, _CENTER, _BORDER

    # 데이터(4행부터)
    r0 = 4
    for i, (label, row) in enumerate(pivot.iterrows()):
        r = r0 + i
        lc = ws.cell(r, 1, label)
        lc.font = _BOLD
        lc.alignment = _LEFT
        lc.border = _BORDER
        rtot = 0
        for j, d in enumerate(dates):
            v = int(row[d])
            rtot += v
            cell = ws.cell(r, 2 + j, v)
            cell.number_format = _NUMFMT
            cell.font = _NORM
            cell.alignment = _CENTER
            cell.border = _BORDER
        tc = ws.cell(r, ncols, rtot)
        tc.number_format = _NUMFMT
        tc.font = _BOLD
        tc.alignment = _CENTER
        tc.border = _BORDER
        tc.fill = _TOTAL
    # 폭·고정 (합계행 3행까지 고정)
    ws.column_dimensions["A"].width = max(12, min(22, max((len(str(x)) for x in pivot.index), default=8) + 4))
    for j, d in enumerate(dates):
        ws.column_dimensions[get_column_letter(2 + j)].width = 7.5 if date_cols else max(9, len(str(d)) + 3)
    ws.column_dimensions[get_column_letter(ncols)].width = 11
    ws.freeze_panes = "A4"
    _fit_page(ws)


def _write_artist_sheet(ws, sub: pd.DataFrame, artist: dict, all_dates, changes_idx=None, name_map=None, country_order=None):
    """아티스트 시트 (RIIZE 참고형): 국가 | 아티스트 | 멤버 | 누적 | 날짜들.
    국가별 블록 + 굵은 구분선 + 국가 소계 + 전체 합계. 날짜는 실제 날짜값."""
    dates = [d for d in all_dates if d in set(sub["날짜"])]
    LAB = 4  # 국가·아티스트·멤버·누적
    ncols = LAB + len(dates)

    # 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    note = " (일본 단독)" if artist["countries"] == ["jp"] else ""
    tc = ws.cell(1, 1, f"{artist['name']} · Artist별 일별 촬영수{note}")
    tc.font = Font(bold=True, size=14, name="맑은 고딕", color="1A1A2E")
    tc.alignment = _LEFT
    ws.row_dimensions[1].height = 24

    # 헤더(2행) — 연한 배경 + 진한 글씨(흰 글씨 안 보이는 문제 방지)
    for j, h in enumerate(["국가", "아티스트", "멤버", "누적"]):
        c = ws.cell(2, 1 + j, h)
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _HEADBORD
    for j, d in enumerate(dates):
        c = ws.cell(2, LAB + 1 + j, _mmdd(d))
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _HEADBORD
    ws.row_dimensions[2].height = 18

    changes_idx = changes_idx or {}
    # 전 국가 로스터(설정 30개국). 판매 없는 국가도 0으로 표기.
    code_name = dict(name_map or {})
    code_name.update(dict(zip(sub["국가코드"].astype(str), sub["국가"])))  # 보강
    # 국가 노출 순서는 전 아티스트 공통(전체 촬영수 기준)으로 통일 — IP마다 순서가 달라지지 않게.
    if artist["countries"]:
        # 국가 제한 IP(예: 재민제노=jp)는 공통 순서를 그 국가로만 필터
        roster = [str(c) for c in artist["countries"]]
        codes = [c for c in (country_order or roster) if str(c) in set(roster)]
        codes += [c for c in roster if str(c) not in set(map(str, codes))]
    elif country_order:
        codes = list(country_order)
    else:
        # 폴백(공통 순서 미전달): 이 아티스트 촬영수 기준
        ctot = sub.groupby("국가코드")["촬영수"].sum()
        codes = sorted(dict.fromkeys(code_name.keys()), key=lambda c: -int(ctot.get(str(c), 0)))
    r = 3
    grand = [0] * len(dates)

    def _cell(row, col, val, *, font=_NORM, fill=None, num=False, left=False, top=_thin, bottom=_thin):
        c = ws.cell(row, col, val)
        c.font = font
        c.alignment = _LEFT if left else _CENTER
        c.border = Border(left=_thin, right=_thin, top=top, bottom=bottom)
        if num:
            c.number_format = _NUMFMT
        if fill:
            c.fill = fill
        return c

    # ── 전체 합계를 상단(헤더 바로 아래)에 — 큰 숫자부터 보이게 ──
    day_sum = sub.groupby("날짜")["촬영수"].sum()
    grand = [int(day_sum.get(d, 0)) for d in dates]
    _cell(r, 1, "전체", font=_HEAD_FONT, fill=_GRAND, left=True, top=_MED, bottom=_MED)
    _cell(r, 2, None, fill=_GRAND, top=_MED, bottom=_MED)
    _cell(r, 3, "합계", font=_HEAD_FONT, fill=_GRAND, left=True, top=_MED, bottom=_MED)
    _cell(r, 4, sum(grand), font=_HEAD_FONT, num=True, fill=_GRAND, top=_MED, bottom=_MED)
    for j, g in enumerate(grand):
        _cell(r, LAB + 1 + j, g, font=_HEAD_FONT, num=True, fill=_GRAND, top=_MED, bottom=_MED)
    r += 1

    for code in codes:
        cc_name = _ctry_label(code, code_name.get(str(code), str(code).upper()))
        pv = pd.pivot_table(sub[sub["국가코드"] == code], index="멤버", columns="날짜",
                            values="촬영수", aggfunc="sum", fill_value=0)
        members = _member_order(artist["name"], set(pv.index))
        csub = [0] * len(dates)
        for i, m in enumerate(members):
            top = _MED if i == 0 else _thin  # 국가 블록 시작에 굵은선
            # 그 국가에서 판매 없던 멤버도 0으로 표기
            vals = [int(pv.loc[m].get(d, 0)) for d in dates] if m in pv.index else [0] * len(dates)
            _cell(r, 1, cc_name if i == 0 else None, font=_BOLD, top=top, left=True)
            _cell(r, 2, artist["name"] if i == 0 else None, top=top)
            _cell(r, 3, m, left=True, top=top)
            _cell(r, 4, sum(vals), font=_BOLD, num=True, top=top)
            for j, (d, v) in enumerate(zip(dates, vals)):
                c = _cell(r, LAB + 1 + j, v, num=True, top=top)
                chg = changes_idx.get((str(d), str(code), artist["name"], m))
                if chg:  # 변동 셀: 빨간 글씨 + 메모
                    prev, new, upd = chg
                    c.font = _CHG_FONT
                    c.comment = Comment(f"이전 {prev:,} → 현재 {new:,}\n({upd} 갱신)", "SM 자동집계")
                csub[j] += v
            r += 1
        # 국가 소계
        _cell(r, 1, None, fill=_TOTAL, bottom=_MED)
        _cell(r, 2, None, fill=_TOTAL, bottom=_MED)
        _cell(r, 3, f"{cc_name} 소계", font=_BOLD, fill=_TOTAL, left=True, bottom=_MED)
        _cell(r, 4, sum(csub), font=_BOLD, num=True, fill=_TOTAL, bottom=_MED)
        for j in range(len(dates)):
            _cell(r, LAB + 1 + j, csub[j], font=_BOLD, num=True, fill=_TOTAL, bottom=_MED)
        r += 1

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 9
    for j in range(len(dates)):
        ws.column_dimensions[get_column_letter(LAB + 1 + j)].width = 6.8
    # 제목·헤더·전체합계(1~3행)까지 고정 — 스크롤해도 합계가 보이게
    ws.freeze_panes = "A4"
    _fit_page(ws)


def _write_changes_sheet(wb, ch: pd.DataFrame, name_map: dict):
    """변경내역 시트: 매주 갱신 시 바뀐 (날짜·국가·멤버) 값 목록(최신순)."""
    ws = wb.create_sheet("변경내역", 1)  # 요약 바로 다음
    heads = ["갱신일", "날짜", "국가", "아티스트", "멤버", "이전", "신규", "증감"]
    for j, h in enumerate(heads):
        c = ws.cell(1, 1 + j, h)
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _HEADBORD
    d = ch.copy()
    d["국가"] = d["국가코드"].map(lambda c: _ctry_label(c, name_map.get(str(c), str(c).upper())))
    d["증감"] = d["신규"].astype(int) - d["이전"].astype(int)
    d = d.sort_values(["갱신일", "증감"], ascending=[False, False])
    r = 2
    for upd, day, ctry, art, mem, prev, new, delta in zip(
            d["갱신일"], d["날짜"], d["국가"], d["아티스트"], d["멤버"],
            d["이전"], d["신규"], d["증감"]):
        row = [str(upd), _mmdd(day), ctry, art, mem, int(prev), int(new), int(delta)]
        for j, v in enumerate(row):
            c = ws.cell(r, 1 + j, v)
            c.border = _BORDER
            c.alignment = _CENTER if j >= 5 else _LEFT
            if j >= 5:
                c.number_format = _NUMFMT
            if j == 7:  # 증감 색
                c.font = Font(name="맑은 고딕", bold=True,
                              color="C0392B" if delta > 0 else "1F6FB2")
        r += 1
    for col, w in zip("ABCDEFGH", (12, 8, 12, 11, 12, 9, 9, 9)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    _fit_page(ws)


def _write_unmatched_sheet(wb, um: pd.DataFrame):
    """미분류 IP(팔리는데 탭 없는 신규/누락) 안내 시트 — 요약 다음."""
    ws = wb.create_sheet("미분류IP", 1)
    ws.merge_cells("A1:E1")
    t = ws.cell(1, 1, "⚠ 팔리고 있지만 아티스트 탭에 없는 IP — 포함하려면 확인 필요")
    t.font = Font(bold=True, size=12, name="맑은 고딕", color="C0392B")
    for j, h in enumerate(["테마(원본)", "프레임수", "최근 7일 촬영", "마지막 판매일", "전체 촬영"]):
        c = ws.cell(2, 1 + j, h)
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _HEADBORD
    r = 3
    for th, fcnt, total, recent, last in zip(
            um["테마"], um["프레임수"], um["총촬영수"], um["최근판매"], um["마지막판매"]):
        for j, v in enumerate([th, int(fcnt), int(recent), str(last), int(total)]):
            c = ws.cell(r, 1 + j, v)
            c.border = _BORDER
            c.alignment = _LEFT if j == 0 else _CENTER
            if j in (1, 2, 4):
                c.number_format = _NUMFMT
        r += 1
    for col, w in zip("ABCDE", (22, 10, 14, 14, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    _fit_page(ws)


def _write_info_sheet(wb, df: pd.DataFrame, a: pd.DataFrame):
    """표지·안내 시트(맨 앞) — 정의·기간·단위·갱신·범례. 되묻기 차단용."""
    ws = wb.create_sheet("안내", 0)
    days = sorted(df["날짜"].astype(str).unique())
    period = f"{days[0]} ~ {days[-1]}" if days else "-"
    artists = [x["name"] for x in ARTISTS if x["name"] in set(a["아티스트"])]
    ws.merge_cells("A1:B1")
    t = ws.cell(1, 1, "SM 아티스트 포토부스 — 일별 촬영수 리포트")
    t.font = Font(bold=True, size=15, name="맑은 고딕", color="1A1A2E")
    rows = [
        ("대상 기간", f"{period}  (각 국가 현지시각 기준 하루)"),
        ("데이터 기준일", days[-1] if days else "-"),
        ("발행일", f"{dt.date.today().isoformat()}  ·  매주 월요일 자동 갱신"),
        ("집계 시각", "각 국가 현지 표준시 00:00~23:59 기준. 시차 국가는 KST와 다릅니다."),
        ("갱신·변동", "최근 2주를 매주 다시 받습니다. 시차로 확정 전 값은 이후 바뀔 수 있어 덮어쓰며, 변동분은 [변경내역] 시트와 셀의 빨간 글씨·메모(이전→현재)로 표시합니다."),
        ("포함 범위", "전 세계 30개국 자체 운영 포토부스."),
        ("오픈 IP", "  ·  ".join(artists)),
    ]
    r = 3
    for label, val in rows:
        lc = ws.cell(r, 1, label)
        lc.font, lc.fill, lc.border = _HEAD_FONT, _HEAD_FILL, _BORDER
        lc.alignment = Alignment(horizontal="left", vertical="top")
        vc = ws.cell(r, 2, val)
        vc.font, vc.border = _NORM, _BORDER
        vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 34 if len(str(val)) > 45 else 18
        r += 1
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 96
    _fit_page(ws)


def _write_member_summary(ws, a: pd.DataFrame):
    """멤버별 요약 — 아티스트·멤버별 총 촬영수 + 판매 국가수(0 멤버도 표기)."""
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, "멤버별 요약 — 총 촬영수 (전 기간 · 전 국가)")
    t.font = Font(bold=True, size=13, name="맑은 고딕", color="1A1A2E")
    for j, h in enumerate(["아티스트", "멤버", "총 촬영수", "판매 국가수"]):
        c = ws.cell(2, 1 + j, h)
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _HEADBORD
    tot = a.groupby(["아티스트", "멤버"])["촬영수"].sum()
    ncc = a[a["촬영수"] > 0].groupby(["아티스트", "멤버"])["국가코드"].nunique()
    present = set(a["아티스트"].unique())
    r = 3
    for art in ARTISTS:
        name = art["name"]
        if name not in present:
            continue
        for m in art["members"].keys():
            vals = [name, m, int(tot.get((name, m), 0)), int(ncc.get((name, m), 0))]
            for j, v in enumerate(vals):
                c = ws.cell(r, 1 + j, v)
                c.border = _BORDER
                c.alignment = _LEFT if j < 2 else _CENTER
                if j >= 2:
                    c.number_format = _NUMFMT
            r += 1
    for col, w in zip("ABCD", (13, 13, 12, 11)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    _fit_page(ws)


def build_xlsx(df: pd.DataFrame) -> bytes:
    a = annotate(df)
    all_dates = sorted(a["날짜"].unique())
    ch = load_changes()
    cidx = _changes_index(ch)
    um = detect_unmatched(df)
    nm = country_name_map()
    wb = Workbook()
    wb.remove(wb.active)

    # 요약: 아티스트 × 날짜 (전 국가 합산)
    summ = pd.pivot_table(a, index="아티스트", columns="날짜", values="촬영수",
                          aggfunc="sum", fill_value=0)
    order = [art["name"] for art in ARTISTS if art["name"] in summ.index]
    ws = wb.create_sheet("요약")
    _write_pivot(ws, summ.reindex(order).astype(int),
                 "SM 촬영 현황 — 아티스트별 일별 합계 (전 국가)", "아티스트")

    # 일자별 요약 — 날짜 × 아티스트
    pv_day = pd.pivot_table(a, index="날짜", columns="아티스트", values="촬영수",
                            aggfunc="sum", fill_value=0).reindex(columns=order, fill_value=0)
    _write_pivot(wb.create_sheet("일자별요약"), pv_day.astype(int),
                 "일자별 요약 — 날짜 × 아티스트 (전 국가)", "날짜", date_cols=False)

    # 국가별 요약 — 국가 × 아티스트 (전 30개국, ISO 병기)
    ac = a.copy()
    ac["국가표기"] = ac["국가코드"].map(lambda c: _ctry_label(c, nm.get(str(c), str(c).upper())))
    pv_ctry = pd.pivot_table(ac, index="국가표기", columns="아티스트", values="촬영수",
                             aggfunc="sum", fill_value=0).reindex(columns=order, fill_value=0)
    pv_ctry = pv_ctry.loc[pv_ctry.sum(axis=1).sort_values(ascending=False).index]  # 촬영수 순
    _write_pivot(wb.create_sheet("국가별요약"), pv_ctry.astype(int),
                 "국가별 요약 — 국가 × 아티스트 (전 기간)", "국가", date_cols=False)

    # 멤버별 요약
    _write_member_summary(wb.create_sheet("멤버별요약"), a)

    # 미분류 IP 안내(있을 때만)
    if not um.empty:
        _write_unmatched_sheet(wb, um)

    # 변경내역(있을 때만, 요약 다음에)
    if not ch.empty:
        _write_changes_sheet(wb, ch, nm)

    # 국가 노출 순서 = 전체 촬영수 내림차순(동률·미판매는 코드순) — 모든 아티스트 시트 공통
    gtot = a.groupby("국가코드")["촬영수"].sum()
    all_codes = set(map(str, nm.keys())) | set(a["국가코드"].astype(str))
    country_order = sorted(all_codes, key=lambda c: (-int(gtot.get(c, 0)), c))

    # 아티스트별 시트: 국가 × 멤버 × 날짜 (변동 셀엔 메모)
    for art in ARTISTS:
        sub = a[a["아티스트"] == art["name"]]
        if sub.empty:
            continue
        ws = wb.create_sheet(art["name"][:31])
        _write_artist_sheet(ws, sub, art, all_dates, cidx, nm, country_order)

    # 표지(맨 앞)
    _write_info_sheet(wb, df, a)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    a = sys.argv[1:]
    start = a[0] if len(a) > 0 else None
    end = a[1] if len(a) > 1 else None
    if not DAILY_PARQUET.exists():
        print("sm_shoot_daily.parquet 가 없어요. 먼저 python sm_collect.py 로 수집하세요.")
        sys.exit(1)
    df = load_daily(start, end)
    if df.empty:
        print("해당 기간 데이터가 없어요.")
        sys.exit(0)
    REPORT_DIR.mkdir(exist_ok=True)
    out = REPORT_DIR / f"SM촬영현황_{df['날짜'].min()}_{df['날짜'].max()}.xlsx"
    out.write_bytes(build_xlsx(df))
    print(f"저장: {out}  (행 {len(df):,} · 국가 {df['국가'].nunique()})")


if __name__ == "__main__":
    main()
