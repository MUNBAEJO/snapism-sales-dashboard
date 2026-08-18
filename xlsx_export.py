# -*- coding: utf-8 -*-
"""표 하나를 **엑셀 파일(.xlsx)** 바이트로. 화면 내려받기 버튼 전용.

왜 CSV 가 아니라 xlsx 인가 (2026-08-12)
    "매출액에 쉼표가 들어가면 좋겠다" 는 요청 때문이다. **CSV 로는 못 한다** —
    `673,958,779` 처럼 쉼표를 박아 넣으면 값 안에 구분자가 생겨 따옴표로 묶이고,
    엑셀은 그걸 **글자로 읽는다.** 보기엔 맞는데 합계·피벗이 안 된다.
    xlsx 는 값은 숫자로 두고 **표시 형식만** `#,##0` 으로 준다 — 보이는 건 쉼표,
    계산은 그대로. 버튼 이름이 '엑셀 다운로드' 인 것과도 맞다.

받는 사람이 바로 쓰도록 몇 가지를 더 해 둔다 — 머리줄 고정·자동 필터·열 너비.
"""
from __future__ import annotations

import io

import pandas as pd

# 열 이름 → 표시 형식. 이름으로 고르므로 새 열이 생겨도 규칙만 맞으면 따라온다.
_MONEY = ("매출", "IP 매출 합계", "건당 평균", "정산금액", "금액")
_COUNT = ("건수", "매장수", "회차수", "판매 국가 수")
_RATE = ("국가 비중(%)", "비중(%)")

_FMT_MONEY = "#,##0"
_FMT_COUNT = "#,##0"
_FMT_RATE = "0.0"
_FMT_DATE = "yyyy-mm-dd"


def _base(col: str) -> str:
    """머리줄에 붙인 단위를 떼고 본 이름만. `매출(원)` → `매출`.
    ★단위를 헤더에 넣기로 하면서(2026-08-12) 형식 매칭이 통째로 깨질 뻔했다."""
    i = col.find("(")
    return col[:i].strip() if i > 0 else col.strip()


def _fmt_of(col: str, series: pd.Series) -> str | None:
    c = _base(col)
    if c in (_base(x) for x in _RATE):
        return _FMT_RATE
    if c in _MONEY:
        return _FMT_MONEY
    if c in _COUNT:
        return _FMT_COUNT
    if pd.api.types.is_datetime64_any_dtype(series):
        return _FMT_DATE
    # 날짜가 object(date) 로 들어오는 경우 — 이름으로 판단한다
    if c.endswith("일") or c.endswith("날짜"):
        return _FMT_DATE
    return None


def _width_of(col: str, series: pd.Series) -> int:
    """한글은 폭이 넓어 글자 수 그대로 쓰면 좁다 → 한글 1.7칸으로 센다."""
    def w(v):
        s = str(v)
        return sum(1.7 if ord(c) > 0x1100 else 1 for c in s)

    head = w(col)
    body = max((w(v) for v in series.head(300)), default=0)
    return int(min(max(head + 2, body + 2, 8), 42))


def _write_sheet(xw, name: str, df: pd.DataFrame, note) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    notes = [n for n in ([note] if isinstance(note, str) else list(note or [])) if n]
    start = len(notes)                # 안내줄 아래부터 표
    name = name[:31]
    df.to_excel(xw, sheet_name=name, index=False, startrow=start)
    ws = xw.sheets[name]
    hdr = start + 1                   # 머리줄의 엑셀 행 번호(1-base)

    for r, n in enumerate(notes, start=1):
        c = ws.cell(row=r, column=1, value=n)
        c.font = Font(size=9, bold=(r == 1), color="2B3350" if r == 1 else "6B7488")

    fill = PatternFill("solid", start_color="EEF1F8")
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=hdr, column=c)
        cell.font = Font(bold=True, color="2B3350")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, col in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = _width_of(col, df[col])
        fmt = _fmt_of(col, df[col])
        if not fmt:
            continue
        for r in range(hdr + 1, hdr + 1 + len(df)):
            ws.cell(row=r, column=i).number_format = fmt

    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)          # 머리줄 고정
    if len(df):
        ws.auto_filter.ref = (f"A{hdr}:"
                              f"{get_column_letter(len(df.columns))}{hdr + len(df)}")


def to_xlsx(df, sheet_name: str = "데이터", note="") -> bytes:
    """DataFrame → xlsx 바이트.

    note 는 문자열 하나 또는 여러 줄(리스트). 표 위에 **조건과 기준**을 적는다 —
    받은 파일만 보고도 "이 숫자가 뭘 어떻게 센 건지" 알 수 있어야 한다.

    ★시트를 여러 장 쓰려면 df 에 {시트이름: 표} 를 넘긴다. note 도 같은 키의
      dict 로 주면 시트마다 다른 안내가 붙는다(문자열 하나면 첫 시트에만).
      기준이 다른 표(원장 vs 테마 리포트)를 한 장에 섞으면 합계가 조용히
      두 번 세진다 — 그래서 장을 나눈다.
    """
    sheets = df if isinstance(df, dict) else {sheet_name: df}
    notes = note if isinstance(note, dict) else {next(iter(sheets)): note}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for nm, one in sheets.items():
            _write_sheet(xw, nm, one, notes.get(nm, ""))
    return buf.getvalue()
